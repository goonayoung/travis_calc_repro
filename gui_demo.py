import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QGroupBox, QRadioButton, QComboBox
, QPushButton, QVBoxLayout, QHBoxLayout, QDesktopWidget, QLineEdit, QLabel, QTableWidget
, QTableWidgetItem, QMessageBox, QTextEdit, QDialog)
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer


class Similarity:

    def listToVector(self, reqList):
        tfidf_vect_simple = TfidfVectorizer()
        reqVector = tfidf_vect_simple.fit_transform(reqList)
        return reqVector
    
    def vectorToDense(self, reqVector):
        denseMatrix = reqVector.todense()
        return denseMatrix
    
    def checkSimilarity(self, denseMatrix):
        vect1 = np.array(denseMatrix[0]).reshape(-1,)
        vect2 = np.array(denseMatrix[1]).reshape(-1,)
        dot_product = np.dot(vect1, vect2)
        l2_norm = (np.sqrt(sum(np.square(vect1))) * np.sqrt(sum(np.square(vect2))))        
        
        if dot_product == 0.0 or l2_norm == 0.0:
            return 0.0
        else:
            similarity = dot_product / l2_norm
            return similarity
    
    def splitWord(self, contentList):
        splitArr,a, b, c, d, e, f, splitWordList = [], [], [], [], [], [], [], []

        for x in range(0, len(contentList), 1):
            splitArr = contentList[x].split("_")
            for y in range(0, len(splitArr), 1):
                a.append(splitArr[y])

        for x in range(0, len(a), 1):
            splitArr = a[x].split(".")
            for y in range(0, len(splitArr), 1):
                b.append(splitArr[y])
        
        for x in range(0, len(b), 1):
            splitArr = b[x].split(" ")
            for y in range(0, len(splitArr), 1):
                c.append(splitArr[y])
        
        for x in range(0, len(c), 1):
            splitArr = c[x].split(":")
            for y in range(0, len(splitArr), 1):
                d.append(splitArr[y])
        
        for x in range(0, len(d), 1):
            splitArr = d[x].split("=")
            for y in range(0, len(splitArr), 1):
                e.append(splitArr[y])
                
        for x in range(0, len(e), 1):
            splitArr = e[x].split("\n")
            for y in range(0, len(splitArr), 1):
                f.append(splitArr[y])
                
        for x in range(0, len(f), 1):
            splitArr = f[x].split("\xa0")
            for y in range(0, len(splitArr), 1):
                splitWordList.append(splitArr[y])
                
        return splitWordList  
    
    def filterKey(self, splitWordList):
        keyList = []
        for x in range(0, len(splitWordList), 1):
            cnt = 0
            a = [ord(c) for c in splitWordList[x]]
            for y in range(0, len(a), 1):
                if(a[y]>47 and a[y]<58):
                    cnt = cnt - 1
            if(splitWordList[x].isupper() == True):
                keyList.append(splitWordList[x])
            elif(cnt < 0):
                keyList.append((splitWordList[x]))
                
        return keyList
                
    def listToStr(self, keyList):
        keyWord = ''
        for x in range(0, len(keyList), 1):
            keyWord = keyWord + keyList[x] + ' '
        return keyWord
        
    def checkREQSimilarity(self, reqArr):
        print("")
        
    def checkTCSimilarity(self, reqArr):
        print("")
        
    def resultSimilarity(self, reqSimilarity, tcSimilarity):
        print("")
        print("<Result Similarity>")
        df = pd.merge(reqSimilarity, tcSimilarity, on = ['Range1', 'Range2', 'Range3', 'Req Index'], how = "outer", right_index = False)
    
        # 입력요구사항-요구사항 유사도 / 입력요구사항 - 테스트케이스 유사도
        # 비교 가중치
        # 가중치 합 = 1
        df['Result Similarity'] = 0.6*df["REQ Similarity"] + 0.4*df["TC Similarity"]
    
        resultSimilarity = df[["Range1", "Range2", "Range3", "Req Index", "TC Index", "Result Similarity"]]
        resultSimilarity = resultSimilarity.sort_values("Result Similarity", ascending = False)
        
        # 엑셀파일로 저장하는 코드
        #resultSimilarity.to_excel('C:/Users/ghj12/Desktop/2020 2학기/소개론/KAI_1123/resultSim.xlsx', sheet_name = 'Sheet1', 
        #    header = True,
        #    index = False, 
        #    startrow = 0, 
        #    startcol = 0, 
        #    )
        
        return resultSimilarity
        

class InternalSimilarity(Similarity):
    def checkREQSimilarity(self, reqArr):
        super().checkREQSimilarity(reqArr)
        df = pd.read_excel(os.getcwd()+'/requirement_internal.xlsx')
        df = df.fillna('000')
        sim=[]
        x=0
        excelAll=df.values
        while(x < len(excelAll)):
            
            arr=excelAll[x][:]
            dataArr = np.ravel(arr, order='C')
            doc = [0, 0]
            simArr = []
            length = len(reqArr)
            for i in range(0, length-1, 1):
                doc[0] = dataArr[i+5]
                doc[1] = reqArr[i]
                reqVector = self.listToVector(doc)
                denseMatrix = self.vectorToDense(reqVector)
                simNum = self.checkSimilarity(denseMatrix)
                simArr.append(simNum)
                
            if(dataArr[8] == reqArr[length-1]):
                simArr.append(1)
            else:
                simArr.append(0)
            
            # Internal) 입력요구사항-요구사항 비교 가중치
            # 가중치 합 = 1
            # Input Source / Signal Description / Type / Packet ID
            weightSum = 0.25*simArr[0] + 0.2*simArr[1] + 0.05*simArr[2] + 0.5*simArr[3]
            
            sim.append([])
            sim[x].append(arr[0])
            sim[x].append(arr[1])
            sim[x].append(arr[2])
            sim[x].append(arr[3])
            sim[x].append(weightSum)
            
            x = x + 1
            
        sim=np.array(sim)
        reqSimilarity = pd.DataFrame(sim, columns= ["Range1", "Range2", "Range3", "Req Index", "REQ Similarity"])
         
        return reqSimilarity
    
    def checkTCSimilarity(self, reqArr):
        super().checkTCSimilarity(reqArr)
        df = pd.read_excel(os.getcwd()+'/internal_testcase.xlsx')
        x = 0
        sim = []
        df = df.fillna('')
        excelAll = df.values
        while(x < len(excelAll) ):
            arr = excelAll[x][:]
            
            contentList = [0,0,0]
            contentList[0]=arr[6]
            contentList[1]=arr[7]
            contentList[2]=arr[8]
            
            splitWordList = self.splitWord(contentList)
            filterKeyList = self.filterKey(splitWordList)
            tcDataToStr = self.listToStr(filterKeyList)
            reqArrToStr = self.listToStr(reqArr)
            
            doc = [0,0]
            doc[0] = tcDataToStr
            doc[1] = reqArrToStr
            reqVector = self.listToVector(doc)
            denseMatrix = self.vectorToDense(reqVector)
            simNum = self.checkSimilarity(denseMatrix)
           
            sim.append([])
            sim[x].append(arr[0])
            sim[x].append(arr[1])
            sim[x].append(arr[2])
            sim[x].append(arr[3])
            sim[x].append(arr[4])
            sim[x].append(simNum)
         
            x = x + 1
         
        sim=np.array(sim)
        tcSimilarity = pd.DataFrame(sim, columns= ["Range1", "Range2", "Range3","Req Index", "TC Index", "TC Similarity"])
        tcSimilarity = tcSimilarity.fillna(0)
         
        return tcSimilarity

    
class ExternalSimilarity(Similarity):
    def checkREQSimilarity(self, reqArr):
        super().checkREQSimilarity(reqArr)
        df = pd.read_excel(os.getcwd()+'/requirement_external.xlsx')
        df = df.fillna('000')
        sim=[]
        x=0
        excelAll=df.values
        while(x < len(excelAll)):
            
            arr=excelAll[x][:]
            dataArr = np.ravel(arr, order='C')
            doc = [0, 0]
            simArr = []
            length = len(reqArr)
            for i in range(0, length, 1):
                doc[0] = dataArr[i+5]
                doc[1] = reqArr[i]
                reqVector = self.listToVector(doc)
                denseMatrix = self.vectorToDense(reqVector)
                simNum = self.checkSimilarity(denseMatrix)
                simArr.append(simNum) 
            
            # External) 입력요구사항-요구사항 비교 가중치
            # 가중치 합 = 1
            # Input Source / Signal Description / ICD Signal Description
            weightSum = 0.3*simArr[0] + 0.2*simArr[1] + 0.5*simArr[2]
            
            sim.append([])
            sim[x].append(arr[0])
            sim[x].append(arr[1])
            sim[x].append(arr[2])
            sim[x].append(arr[3])
            sim[x].append(weightSum)
            
            x = x + 1
            
            
        sim=np.array(sim)
        reqSimilarity = pd.DataFrame(sim, columns= ["Range1", "Range2", "Range3", "Req Index", "REQ Similarity"])

        return reqSimilarity
    
    def checkTCSimilarity(self, reqArr):
        super().checkTCSimilarity(reqArr)
        df = pd.read_excel(os.getcwd()+'/external_testcase.xlsx')
        x = 0
        sim = []
        df = df.fillna('')
        excelAll = df.values
        while(x < len(excelAll) ):
            arr = excelAll[x][:]
            
            contentList = [0,0,0]
            contentList[0]=arr[6]
            contentList[1]=arr[7]
            contentList[2]=arr[8]
            splitWordList = self.splitWord(contentList)
            filterKeyList = self.filterKey(splitWordList)
            tcDataToStr = self.listToStr(filterKeyList)
            reqArrToStr = self.listToStr(reqArr)
            
            doc = [0,0]
            doc[0] = tcDataToStr
            doc[1] = reqArrToStr
            reqVector = self.listToVector(doc)
            denseMatrix = self.vectorToDense(reqVector)
            simNum = self.checkSimilarity(denseMatrix)
           
            sim.append([])
            sim[x].append(arr[0])
            sim[x].append(arr[1])
            sim[x].append(arr[2])
            sim[x].append(arr[3])
            sim[x].append(arr[4])
            sim[x].append(simNum)
         
            x = x + 1
         
        sim=np.array(sim)
        tcSimilarity = pd.DataFrame(sim, columns= ["Range1", "Range2", "Range3","Req Index", "TC Index", "TC Similarity"])
        tcSimilarity = tcSimilarity.fillna(0)
         
        return tcSimilarity



class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        vbox = QVBoxLayout()
        vbox.addWidget(self.Setting())
        vbox.addWidget(self.Input())
        vbox.addWidget(self.Output())

        self.setLayout(vbox)

        self.setWindowTitle('KAI Testcase Recommandation Program')
        self.resize(800, 800)
        self.center()
        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def Setting(self):
        groupbox = QGroupBox('Setting')

        self.groupbox1 = QGroupBox('검색 범위')
        self.groupbox1.setFlat(True)
        self.groupbox2 = QGroupBox('결과 갯수')
        self.groupbox2.setFlat(True)
        self.groupbox1.setCheckable(True)
        self.groupbox1.setChecked(False)
        self.groupbox2.setCheckable(True)
        self.groupbox2.setChecked(False)

        self.cb1 = QComboBox(self)
        self.cb1.addItem('3.2.11')
        self.cb1.addItem('3.2.12')
        self.cb1.addItem('3.2.13')
        self.cb1.addItem('3.2.14')
        label1 = QLabel('~',self)
        self.cb2 = QComboBox(self)
        self.cb2.addItem('3.2.11')
        self.cb2.addItem('3.2.12')
        self.cb2.addItem('3.2.13')
        self.cb2.addItem('3.2.14')
        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.cb1)
        hbox1.addWidget(label1)
        hbox1.addWidget(self.cb2)
        self.groupbox1.setLayout(hbox1)
        
        self.cb2.setCurrentText('3.2.14')
        
        self.SearchR11 = int(self.cb1.currentText()[0])
        self.SearchR12 = int(self.cb1.currentText()[2])
        self.SearchR13 = int(self.cb1.currentText()[4:])
        self.SearchR21 = 3
        self.SearchR22 = 2
        self.SearchR23 = 14

        self.qle3 = QLineEdit(self)
        self.qle3.setText('1')
        label2 = QLabel('개',self)
        
        self.outputnum = int(self.qle3.text())
        
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.qle3)
        hbox2.addWidget(label2)
        self.groupbox2.setLayout(hbox2)
        
        self.setBtn = QPushButton('Set',self)
        self.setBtn.setCheckable(True)
        self.setBtn.clicked.connect(self.Set_clicked)

        hbox = QHBoxLayout()
        hbox.addWidget(self.groupbox1)
        hbox.addWidget(self.groupbox2)
        hbox.addStretch(1)
        hbox.addWidget(self.setBtn)
        groupbox.setLayout(hbox)
        

        return groupbox

    def Set_clicked(self):       
        if(self.groupbox1.isChecked()):
            self.cb1.currentTextChanged
            self.cb2.currentTextChanged
            self.SearchR11 = int(self.cb1.currentText()[0])
            self.SearchR12 = int(self.cb1.currentText()[2])
            self.SearchR13 = int(self.cb1.currentText()[4:])
            self.SearchR21 = int(self.cb2.currentText()[0])
            self.SearchR22 = int(self.cb2.currentText()[2])
            self.SearchR23 = int(self.cb2.currentText()[4:])

            if(self.SearchR13>self.SearchR23):
                print("범위를 다시 설정하세요.")
                QMessageBox.about(self, '범위설정 오류', '범위를 다시 설정해주세요.')
            else:
                print("Search Range1 is : " ,self.SearchR11)
                print("Search Range1 is : " ,self.SearchR12)
                print("Search Range1 is : " ,self.SearchR13)
                print("Search Range2 is : ", self.SearchR21)
                print("Search Range2 is : ", self.SearchR22)
                print("Search Range2 is : ", self.SearchR23)
        else:
            self.SearchR11 = 3
            self.SearchR12 = 2
            self.SearchR13 = 11
            self.SearchR21 = 3
            self.SearchR22 = 2
            self.SearchR23 = 14
            
            print("Search Range1 is : " ,self.SearchR11)
            print("Search Range1 is : " ,self.SearchR12)
            print("Search Range1 is : " ,self.SearchR13)
            print("Search Range2 is : ", self.SearchR21)
            print("Search Range2 is : ", self.SearchR22)
            print("Search Range2 is : ", self.SearchR23)
            
        
        if(self.groupbox2.isChecked()):
            try:
                self.outputnum = int(self.qle3.text())
                print("outputnum is : " , self.outputnum)
            except:
                QMessageBox.about(self, '결과개수오류','숫자를 입력해 주세요.')
        else:
            self.qle3.setText('1')
            self.outputnum = int(self.qle3.text())
            print("outputnum is : " , self.outputnum)
        
        print('Set')


    def Input(self):
        groupbox = QGroupBox('Input')

        hbox1 = QHBoxLayout()
        self.I_button = QRadioButton('Internal Input')
        self.E_button = QRadioButton('External Input')
        self.I_button.setChecked(True)
        self.I_button.clicked.connect(self.I_clicked)
        self.E_button.clicked.connect(self.E_clicked)
        hbox1.addWidget(self.I_button)
        hbox1.addWidget(self.E_button)

        Tableinputbox = QHBoxLayout()
        self.table = QTableWidget()
        Tableinputbox.addWidget(self.table)

        self.table.clear()
        self.table.setFixedHeight(90)
        self.table.setRowCount(1)
        self.table.setColumnCount(4)
        column_headers = ['INPUT SOURCE', 'SIGNAL DESCRIPTION', 'TYPE', 'PACKET ID']
        self.table.setHorizontalHeaderLabels(column_headers)
        self.table.setColumnWidth(1, 361)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 80)

        hbox3 = QHBoxLayout()
        self.searchBtn = QPushButton('Search')
        self.searchBtn.clicked.connect(self.Search_clicked)
        hbox3.addStretch(2)
        hbox3.addWidget(self.searchBtn)

        vbox = QVBoxLayout()
        vbox.addLayout(hbox1)
        vbox.addLayout(Tableinputbox)
        vbox.addLayout(hbox3)

        groupbox.setLayout(vbox)
        groupbox.setFixedHeight(200)

        return groupbox

    def I_clicked(self):
        print('internal')
        self.table.clear()
        self.table.setFixedHeight(90)
        self.table.setRowCount(1)
        self.table.setColumnCount(4)
        column_headers = ['INPUT SOURCE', 'SIGNAL DESCRIPTION', 'TYPE', 'PACKET ID']
        self.table.setHorizontalHeaderLabels(column_headers)
        self.table.setColumnWidth(1, 361)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 80)

    def E_clicked(self):
        print('external')
        self.table.clear()
        self.table.setFixedHeight(90)
        self.table.setRowCount(1)
        self.table.setColumnCount(3)
        column_headers = ['INPUT SOURCE', 'SIGNAL DESCRIPTION', 'ICD SIGNAL SPECIFICATION']
        self.table.setHorizontalHeaderLabels(column_headers)
        self.table.setColumnWidth(1, 361)
        self.table.setColumnWidth(2, 160)
        
        
    def Search_clicked(self):
        print("Search")
        self.moreCnt = 0
        self.loopCnt = 0;
        outputnum = self.outputnum
        row = self.table.rowCount()
        column = self.table.columnCount()
        
                    
        inputData = []
        for x in range(0, row, 1):
            for y in range(0, column, 1):
                if self.table.item(x,y) is None:
                    text = ''
                else:
                    text = self.table.item(x, y).text()
                inputData.append(text)

        print(inputData)
        self.testInput1 = inputData[0]
        self.testInput2 = inputData[1]
        self.testInput3 = inputData[2]

        if (inputData == ['','','',''] or inputData == ['', '', '']):
            self.inputIsEmpty()
        else:
            self.outputTable.clearContents()

            if (column == 3):
                similarityClass = ExternalSimilarity()
                tcdf = pd.read_excel(os.getcwd()+'/external_testcase.xlsx')
            elif (column == 4):
                similarityClass = InternalSimilarity()
                tcdf = pd.read_excel(os.getcwd()+'/internal_testcase.xlsx')
                self.testInput4 = inputData[3]
                
            self.outputTable.setRowCount(1)
            self.outputTable.setColumnCount(3)
            

            reqSim = similarityClass.checkREQSimilarity(inputData)
            tcSim = similarityClass.checkTCSimilarity(inputData)
            resultSim = similarityClass.resultSimilarity(reqSim, tcSim)

            df2 = pd.merge(resultSim, tcdf, on=['Range1', 'Range2', 'Range3',
                                            'Req Index', 'TC Index'],
                       how="outer", right_index=False)
            df2 = df2.fillna("-")
            print("콜럼명 : ", list(df2))
            self.arr = df2.values
            self.x = 0;
            self.rangedTC = []
            
            for x in range(0, len(df2.index), 1):
                arr2 = self.arr[x][:]
                if(arr2[0]>=self.SearchR11 and arr2[1]>=self.SearchR12 and arr2[2]>=self.SearchR13 and arr2[0]<=self.SearchR21 and arr2[1]<=self.SearchR22 and arr2[2]<= self.SearchR23):
                    self.rangedTC.append(arr2)
            print("arr3 = ", self.rangedTC[0][7])
            
            if(self.outputnum > len(self.rangedTC)):
                outputnum = len(self.rangedTC)
                    
            for self.x in range(0, outputnum, 1):
                TestAction = self.rangedTC[self.x][7]
                ExpectedResult = self.rangedTC[self.x][8]
                PassFail = self.rangedTC[self.x][9]
                rowPosition = self.outputTable.rowCount()
                if(rowPosition < outputnum):
                    self.outputTable.insertRow(rowPosition)
                self.outputTable.setItem(self.x, 0, QTableWidgetItem(TestAction))
                self.outputTable.setItem(self.x, 1, QTableWidgetItem(ExpectedResult))
                self.outputTable.setItem(self.x, 2, QTableWidgetItem(PassFail))
                   

    def inputIsEmpty(self):
        QMessageBox.warning(self, '경고', '입력 table에 값을 입력해주십시오.',)

    def Output(self):
        groupbox = QGroupBox('Output')

        vbox = QVBoxLayout()
        vbox.addWidget(self.OuputTable())

        hbox = QHBoxLayout()
        self.moreBtn = QPushButton('More')
        self.saveBtn = QPushButton('Save')
        self.moreBtn.clicked.connect(self.More_clicked)
        self.saveBtn.clicked.connect(self.Save_clicked) 
        hbox.addStretch(2)
        hbox.addWidget(self.moreBtn)
        hbox.addWidget(self.saveBtn)

        vbox.addLayout(hbox)

        groupbox.setLayout(vbox)

        return groupbox

    def More_clicked(self):
        print("more")
        self.moreCnt = self.moreCnt + 1  
        self.totalCnt = self.outputnum + self.moreCnt - 1
        
        if(self.totalCnt >= len(self.rangedTC)):
            QMessageBox.warning(self, '경고', '모든 testcase를 표시하였습니다.',)
            return 0
        
        TestAction = self.rangedTC[self.totalCnt][7]
        ExpectedResult = self.rangedTC[self.totalCnt][8]
        PassFail = self.rangedTC[self.totalCnt][9]

        rowPosition = self.outputTable.rowCount()
            
        
        self.outputTable.insertRow(rowPosition)
        self.outputTable.setItem(self.totalCnt, 0, QTableWidgetItem(TestAction))
        self.outputTable.setItem(self.totalCnt, 1, QTableWidgetItem(ExpectedResult))
        self.outputTable.setItem(self.totalCnt, 2, QTableWidgetItem(PassFail))    
        
    def Save_clicked(self):

        print("save")
        TestAction = []
        ExpectedResult = []
        PassFail = []

        for x in range(0, self.totalCnt+1, 1):            
            text1 = self.outputTable.item(x, 0).text() 
            text2 = self.outputTable.item(x, 1).text()
            text3 = self.outputTable.item(x, 2).text()
            TestAction.append(text1)
            ExpectedResult.append(text2)
            PassFail.append(text3)         
        print("TestAction = ", TestAction)
        print("Expected Result = ", ExpectedResult)
        print("PAssFail = ", PassFail)
        df = pd.DataFrame({'Test Action': TestAction, 'Expected Result': ExpectedResult, 'Pass/Fail' : PassFail})       
        print("df = ", df)
        SaveWindow(self, df)

    def OuputTable(self):
        box = QGroupBox()

        self.outputTable = QTableWidget()
        self.outputTable.setRowCount(1)
        self.outputTable.setColumnCount(3)
        column_headers = ['Test Action', 'Expected Result', 'Pass/Fail']
        self.outputTable.setHorizontalHeaderLabels(column_headers)
        self.outputTable.setColumnWidth(0, 260)
        self.outputTable.setColumnWidth(1, 281)
        self.outputTable.setColumnWidth(2, 80)
        vbox = QVBoxLayout()
        vbox.addWidget(self.outputTable)
        box.setLayout(vbox)

        return box

class SaveWindow(QDialog):
    def __init__(self,parent, df):
        self.df = df
        super(SaveWindow,self).__init__(parent)
        self.save()
        self.show()
        
    def save(self):
        lbl1 = QLabel('저장할 파일 경로')
        self.qle1 = QLineEdit()
        self.qle1.setText(os.getcwd())

        lbl2 = QLabel('저장할 파일 이름')
        self.qle2 = QLineEdit()

        btn = QPushButton('파일 생성')
        btn.clicked.connect(self.saveFile_btn)

        vbox = QVBoxLayout()
        vbox.addWidget(lbl1)
        vbox.addWidget(self.qle1)
        vbox.addWidget(lbl2)
        vbox.addWidget(self.qle2)
        vbox.addWidget(btn)

        self.setLayout(vbox)

        self.setWindowTitle('Save')
        self.setGeometry(100, 100, 250, 300)
        
        
    def saveFile_btn(self):
        
        if len(self.qle1.text()) == 0 or len(self.qle2.text()) == 0 :
            return
        directory = self.qle1.text() + '/'
        if os.path.isdir(directory):
            print('적합한 주소입니다')
        else:
            print('파일 경로가 존재하지 않습니다.')
            QMessageBox.about(self, "message", "적합하지 않은 주소입니다")
            return
        wb = openpyxl.Workbook()
        name = self.qle2.text() + '.xlsx'
        
        address = directory+name
        ws=wb.create_sheet('Sheet 1')
        ws=wb.active
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws.append(r)
        wb.save(address)
        QMessageBox.about(self, "message", "파일이 저장되었습니다")
        
        self.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())